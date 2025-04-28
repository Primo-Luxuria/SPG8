import csv
import urllib.parse
import xml.etree.ElementTree as ET
import zipfile
from http.client import responses
from datetime import date
import openpyxl
from django.db.models.fields import *
from openpyxl.utils import get_column_letter
from django.db import connection, IntegrityError
import copy
import pandas as pd
import io
from django.apps import apps
from django.db import models

from bs4 import BeautifulSoup
from django.core.files.base import ContentFile
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.middleware.csrf import get_token
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import login_required

from welcome.models import *

import json
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from .models import Course#, AnswerOption  # Import additional models as needed


def home(request):
    return render(request, 'welcome/home.html')

def faq_view(request):
    return render(request, 'welcome/faq.html')  # or whatever template you want

def signup_handler(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        passwordconfirm = request.POST.get("passwordconfirm")
        role = request.POST.get("role")

        # Check if passwords match
        if password != passwordconfirm:
            messages.error(request, "Passwords do not match.")
            return redirect("signup")  # Changed to redirect to view, not handler

        # Check if the username already exists
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already taken.")
            return redirect("signup")  # Changed to redirect to view, not handler

        # Create new user
        user = User.objects.create_user(username=username, password=password)

        # Create and save user profile with the selected role
        user_profile = UserProfile(user=user, role=role)
        user_profile.save()  # Save the profile after setting the role

        messages.success(request, "Account created successfully!")

        # Redirect based on role
        if role == "teacher":
            return redirect("teacher_dashboard")
        elif role == "publisher":
            return redirect("publisher_dashboard")
        else:
            return redirect("webmaster_dashboard")  # Add a case for 'webmaster' if needed

    return render(request, "welcome/signup.html")



def login_handler(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        if not username or not password:
            messages.error(request, "Username and password are required.")
            return redirect("login")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)  # Log the user in
            
            # Get role from UserProfile
            try:
                role = user.userprofile.role
            except:
                role = "user"  # Default fallback
            
            # Redirect based on role
            if role == "teacher":
                return redirect("teacher_dashboard")
            elif role == "publisher":
                return redirect("publisher_dashboard")
            elif role == "webmaster":
                return redirect("webmaster_dashboard")
            else:
                return redirect("home")  # Default fallback
        else:
            messages.error(request, "Invalid username or password")
            return redirect("login")
    
    return redirect("login")  # Redirect if not POST


import time
def parse_qti_xml(request):
    """
    Parses a QTI XML file and saves extracted data to the database.
    This supports QTI version 1.2 only.
    """
    start_time = time.perf_counter()
    
    class ImageDataPair:
        def __init__(self, raw_image_data, actual_image_name):
            self.raw_image_data = raw_image_data
            self.actual_image_name = actual_image_name

    # Function to remove namespaces
    def remove_namespace(given_tree):
        for elem in given_tree.iter():
            if "}" in elem.tag:
                elem.tag = elem.tag.split("}")[-1]

    # creates a new question record/entry
    def create_question(g_course, g_q_type, g_q_text, g_points):
        temp_question_instance = Question.objects.create(
            course=g_course,
            # this is because, logically, when questions/tests are uploaded to a course, are they not part of it?
            qtype=g_q_type,
            text=g_q_text,
            score=g_points
        )

        # checks if user is logged in
        if request.user.is_authenticated:
            temp_question_instance.author = request.user  # sets to the current user
            temp_question_instance.save()

        return temp_question_instance

    def check_embedded_graphic(text_q):
        if text_q is None:
            return None
        # Parse the HTML using BeautifulSoup4 library
        soup = BeautifulSoup(text_q, 'html.parser')
        # Find the <img> tag
        img_element = soup.find('img')
        found_the_image = False
        if img_element:  # if we find an embedded image
            temp_file_path1 = img_element.get('src')  # gets the src attribute of img element
            url_encoded_path = temp_file_path1[18:]  # length of "$IMS-CC-FILEBASE$/" is 18
            decoded_path = urllib.parse.unquote(url_encoded_path)  # this decodes URL-encoded path

            for potential_image_name in filename_list:  # this assumes an outer function has: filename_list = zip_ref.namelist()
                if potential_image_name.endswith(decoded_path):
                    found_the_image = True
                    my_image_name = img_element.get('alt')
                    with zip_ref.open(potential_image_name) as desired_img_file:  # open the image file
                        img_data = desired_img_file.read()  # this is the raw image data
                        data_to_return = ImageDataPair(img_data, my_image_name)
            if not found_the_image:
                print('Desired image not found')  # used for debugging
        if img_element is None or found_the_image == False:
            return None
        else:
            return data_to_return


    def field_exists(given_model_name, given_field_name):

        try:
            model = apps.get_model(app_label='welcome', model_name=given_model_name)  # get Model class from Welcome app in Django
        except LookupError:  # if model could not be found
            print(f'Model {given_model_name} not found.')
            return False

        model_fields_names_list = [field.name for field in model._meta.fields]  # model._meta.fields is a list of field objects

        if given_field_name in model_fields_names_list:
            return True
        else:
            return False

    def parse_just_xml(meta_path, non_meta_path, the_course, template_instance):

        print(f"processing file: {meta_path}")
        # path to metadata file
        xml_file_path = meta_path
        tree = ET.parse(xml_file_path)
        root = tree.getroot()
        remove_namespace(root)

        cover_instructions_text = root.find('.//description').text

        print(f"processing file: {non_meta_path}")
        # Path to the questions file
        xml_file_path = non_meta_path
        tree = ET.parse(xml_file_path)
        root = tree.getroot()

        namespace = {"ns": "http://www.imsglobal.org/xsd/ims_qtiasiv1p2"}

        # Remove namespaces
        remove_namespace(root)
        # Find the 'assessment' element
        my_tag = "assessment"
        node = root.find(my_tag)

        if node is None:
            return JsonResponse({"error": f"Element '{my_tag}' not found in XML!"}, status=400)

        # Extract 'ident' and 'title' attribute from the element that node represents
        the_test_title = node.get("title")  # test name
        test_identifier = node.get("ident")

        # Create a new Test record
        test_instance = Test.objects.create(
            course=the_course,
            name=the_test_title,
            template=template_instance,
            is_final=True
        )
        if field_exists('Test', 'templateID'):
            test_instance.templateID = template_instance.id
            test_instance.save()

        test_part_instance = TestPart.objects.create(
            test=test_instance
        )
        number_of_sections = 0

        mc_item_list = []
        tf_item_list = []
        fb_item_list = []
        es_item_list = []
        ma_item_list = []
        ms_item_list = []

        for section in root.findall(".//section"):
            # number_of_sections = number_of_sections + 1

            for item in section.findall(".//item"):

                # all useful metadata fields are found in the fieldentry elements under itemmetadata
                node = item.find('itemmetadata')
                qti_metadata_fields = node.findall(".//fieldentry")

                node = item.find('presentation')

                temp_node = node.find('material')  # possibly redundant statement
                temp_node = temp_node.find(".//mattext")
                # question_text_field contains the question prompt text
                question_text_field = temp_node.text

                # each response has an ID represented as "ident" in the XML
                # the ID is used to know what the correct response is
                correct_answer_ident = None

                the_question_type = qti_metadata_fields[0].text
                max_points_for_question = float(qti_metadata_fields[1].text)

                # node should currently be already = element w 'presentation' tag
                # MultiChoice & TF might be able to be combined. For now, they are separate
                if the_question_type == 'multiple_choice_question':
                    #

                    the_question_type = 'mc'

                    node = node.find('.//response_lid')
                    answer_choices_dict = {}
                    for response_label_elem in node.findall('.//response_label'):
                        response_ident = response_label_elem.get('ident')
                        response_text = response_label_elem.find('.//mattext').text
                        answer_choices_dict[response_ident] = response_text

                    node = item.find('resprocessing')
                    for respcondition_elem in node.findall('.//respcondition'):
                        if respcondition_elem.get('continue') == "No":
                            temp_node = respcondition_elem.find('.//varequal')
                            correct_answer_ident = temp_node.text

                    # this creates a question record in database
                    question_instance = create_question(the_course, the_question_type, question_text_field,
                                                        max_points_for_question)
                    # this block of code checks for an embedded graphic in text, then saves it to field
                    # is there is a graphic. otherwise it does nothing.
                    image_data_pair = check_embedded_graphic(question_text_field)
                    if image_data_pair is not None:
                        # Save the image to the img field, then update the record/entry
                        question_instance.img.save(image_data_pair.actual_image_name,
                                                   ContentFile(image_data_pair.raw_image_data))
                        question_instance.save()
                        print(f"{question_instance.img.url}")

                        # Parse the HTML using BeautifulSoup4 library
                        html_obj = BeautifulSoup(question_text_field, 'html.parser')
                        # Find the element with "img" tag
                        my_img_element = html_obj.find('img')
                        my_img_element['src'] = question_instance.img.url  # change src attribute
                        question_text_field = str(html_obj)  # save html as string
                        question_instance.text = question_text_field  # update text field
                        question_instance.save()  # save/update entry in database

                    """
                    testquestion_instance = TestQuestion.objects.create(
                        test=test_instance,
                        question=question_instance,
                        assigned_points=max_points_for_question,
                        section=test_section_instance
                    )
                    """

                    mc_item_list.append({'question': question_instance, 'assigned_points': max_points_for_question})

                    current_place = 0
                    for key, value in answer_choices_dict.items():
                        current_place = current_place + 1
                        if key == correct_answer_ident:
                            #question_instance.answer = value
                            option_letter = get_column_letter(current_place)
                            answer_instance = Answers.objects.create(
                                text=option_letter,
                                question=question_instance
                            )
                            option_instance = Options.objects.create(
                                text=value,
                                question=question_instance,
                                order=current_place
                            )
                            temp_img_data_pair = check_embedded_graphic(value)
                            if temp_img_data_pair is not None:

                                # Parse the HTML using BeautifulSoup4 library
                                html_obj = BeautifulSoup(value, 'html.parser')
                                # Find the element with "img" tag
                                my_img_element = html_obj.find('img')

                                if field_exists('Question', 'ansimg'):
                                    question_instance.ansimg.save(temp_img_data_pair.actual_image_name,
                                                                  ContentFile(temp_img_data_pair.raw_image_data))
                                    my_img_element['src'] = question_instance.ansimg.url  # change src attribute

                                    value = str(html_obj)  # save html as string

                                    if field_exists('Question', 'answer'):
                                        question_instance.answer = value  # update answer text field

                                    question_instance.save()  # save/update entry in database

                                if field_exists('Answers', 'answer_graphic'):
                                    answer_instance.answer_graphic.save(temp_img_data_pair.actual_image_name,
                                                                        ContentFile(temp_img_data_pair.raw_image_data))
                                    my_img_element['src'] = answer_instance.answer_graphic.url  # change src attribute

                                    value = str(html_obj)  # save html as string

                                    option_instance.text = value
                                    option_instance.save()

                                if field_exists('Options', 'image'):
                                    option_instance.image.save(temp_img_data_pair.actual_image_name,
                                                                        ContentFile(temp_img_data_pair.raw_image_data))
                                    my_img_element['src'] = option_instance.image.url

                                    value = str(html_obj)  # save html as string

                                    option_instance.text = value
                                    option_instance.save()

                            question_instance.save()
                        else:
                            options_instance = Options.objects.create(
                                text=value,
                                question=question_instance,
                                order=current_place
                            )
                            temp_img_data_pair = check_embedded_graphic(value)
                            if temp_img_data_pair is not None:
                                options_instance.image.save(temp_img_data_pair.actual_image_name,
                                                            ContentFile(temp_img_data_pair.raw_image_data))
                                options_instance.save()

                                # Parse the HTML using BeautifulSoup4 library
                                html_obj = BeautifulSoup(value, 'html.parser')
                                # Find the element with "img" tag
                                my_img_element = html_obj.find('img')
                                my_img_element['src'] = options_instance.image.url  # change src attribute
                                value = str(html_obj)  # save html as string
                                options_instance.text = value  # update options text field
                                options_instance.save()  # save/update entry in database

                elif the_question_type == 'true_false_question':
                    the_question_type = 'tf'

                    node = node.find('.//response_lid')
                    answer_choices_dict = {}
                    for response_label_elem in node.findall('.//response_label'):
                        response_ident = response_label_elem.get('ident')
                        response_text = response_label_elem.find('.//mattext').text
                        answer_choices_dict[response_ident] = response_text

                    node = item.find('resprocessing')
                    for respcondition_elem in node.findall('.//respcondition'):
                        if respcondition_elem.get('continue') == "No":
                            temp_node = respcondition_elem.find('.//varequal')
                            correct_answer_ident = temp_node.text

                    question_instance = create_question(the_course, the_question_type, question_text_field,
                                                        max_points_for_question)
                    # this block of code checks for an embedded graphic in text, then saves it to field
                    # is there is a graphic. otherwise it does nothing.
                    image_data_pair = check_embedded_graphic(question_text_field)
                    if image_data_pair is not None:
                        # Save the image to the img field, then update the record/entry
                        question_instance.img.save(image_data_pair.actual_image_name,
                                                   ContentFile(image_data_pair.raw_image_data))
                        question_instance.save()
                        print(f"{question_instance.img.url}")

                        # Parse the HTML using BeautifulSoup4 library
                        html_obj = BeautifulSoup(question_text_field, 'html.parser')
                        # Find the element with "img" tag
                        my_img_element = html_obj.find('img')
                        my_img_element['src'] = question_instance.img.url  # change src attribute
                        question_text_field = str(html_obj)  # save html as string
                        question_instance.text = question_text_field  # update text field
                        question_instance.save()  # save/update entry in database

                    """
                    testquestion_instance = TestQuestion.objects.create(
                        test=test_instance,
                        question=question_instance,
                        assigned_points=max_points_for_question,
                        section=test_section_instance
                    )
                    """

                    tf_item_list.append({'question': question_instance, 'assigned_points': max_points_for_question})

                    current_place = 0
                    for key, value in answer_choices_dict.items():
                        current_place = current_place + 1
                        if key == correct_answer_ident:
                            if field_exists('Question', 'answer'):
                                question_instance.answer = value
                            answer_instance = Answers.objects.create(
                                text=value.lower(),
                                question=question_instance
                            )

                    current_place = 0
                    for key, value in answer_choices_dict.items():
                        current_place = current_place + 1
                        option_instance = Options.objects.create(
                            text=value,
                            question=question_instance,
                            order=current_place
                        )

                elif the_question_type == 'short_answer_question':  # fill-in-the-blank question (single)

                    the_question_type = "fill_in_the_blank"

                    the_question_type = 'fb'

                    question_instance = create_question(
                        the_course, the_question_type, question_text_field,
                        max_points_for_question
                    )
                    # this block of code checks for an embedded graphic in text, then saves it to field
                    # is there is a graphic. otherwise it does nothing.
                    image_data_pair = check_embedded_graphic(question_text_field)
                    if image_data_pair is not None:
                        # Save the image to the img field, then update the record/entry
                        question_instance.img.save(image_data_pair.actual_image_name,
                                                   ContentFile(image_data_pair.raw_image_data))
                        question_instance.save()
                        print(f"{question_instance.img.url}")

                        # Parse the HTML using BeautifulSoup4 library
                        html_obj = BeautifulSoup(question_text_field, 'html.parser')
                        # Find the element with "img" tag
                        my_img_element = html_obj.find('img')
                        my_img_element['src'] = question_instance.img.url  # change src attribute
                        question_text_field = str(html_obj)  # save html as string
                        question_instance.text = question_text_field  # update text field
                        question_instance.save()  # save/update entry in database

                    """
                    testquestion_instance = TestQuestion.objects.create(
                        test=test_instance,
                        question=question_instance,
                        assigned_points=max_points_for_question,
                        section=test_section_instance
                    )
                    """

                    fb_item_list.append({'question': question_instance, 'assigned_points': max_points_for_question})

                    node = item.find('resprocessing')
                    for respcondition_elem in node.findall('.//respcondition'):
                        if respcondition_elem.get('continue') == "No":
                            for varequal_elem in respcondition_elem.findall('.//varequal'):
                                answer_instance = Answers.objects.create(
                                    question=question_instance,
                                    text=varequal_elem.text
                                )
                                temp_img_data_pair = check_embedded_graphic(varequal_elem.text)
                                if temp_img_data_pair is not None:
                                    answer_instance.answer_graphic.save(temp_img_data_pair.actual_image_name,
                                                                        ContentFile(temp_img_data_pair.raw_image_data))
                                    answer_instance.save()

                elif the_question_type == 'multiple_answers_question':  # Multiple Selections question

                    the_question_type = 'ms'

                    correct_answer_ident_list = []
                    node = node.find('.//response_lid')
                    answer_choices_dict = {}
                    for response_label_elem in node.findall('.//response_label'):
                        response_ident = response_label_elem.get('ident')
                        response_text = response_label_elem.find('.//mattext').text
                        answer_choices_dict[response_ident] = response_text

                    node = item.find('resprocessing')
                    for respcondition_elem in node.findall('.//respcondition'):
                        if respcondition_elem.get('continue') == "No":
                            for varequal_elem in respcondition_elem.find('conditionvar').find('and').findall(
                                    'varequal'):
                                correct_answer_ident_list.append(varequal_elem.text)

                    question_instance = create_question(the_course, the_question_type, question_text_field,
                                                        max_points_for_question)
                    # this block of code checks for an embedded graphic in text, then saves it to field
                    # is there is a graphic. otherwise it does nothing.
                    image_data_pair = check_embedded_graphic(question_text_field)
                    if image_data_pair is not None:
                        # Save the image to the img field, then update the record/entry
                        question_instance.img.save(image_data_pair.actual_image_name,
                                                   ContentFile(image_data_pair.raw_image_data))
                        question_instance.save()
                        print(f"{question_instance.img.url}")

                        # Parse the HTML using BeautifulSoup4 library
                        html_obj = BeautifulSoup(question_text_field, 'html.parser')
                        # Find the element with "img" tag
                        my_img_element = html_obj.find('img')
                        my_img_element['src'] = question_instance.img.url  # change src attribute
                        question_text_field = str(html_obj)  # save html as string
                        question_instance.text = question_text_field  # update text field
                        question_instance.save()  # save/update entry in database

                    """
                    testquestion_instance = TestQuestion.objects.create(
                        test=test_instance,
                        question=question_instance,
                        assigned_points=max_points_for_question,
                        section=test_section_instance
                    )
                    """

                    ms_item_list.append({'question': question_instance, 'assigned_points': max_points_for_question})

                    current_place = 0
                    for key, value in answer_choices_dict.items():
                        current_place = current_place + 1
                        if key in correct_answer_ident_list:
                            answer_instance = Answers.objects.create(
                                question=question_instance,
                                text=value
                            )
                            option_instance = Options.objects.create(
                                text=value,
                                question=question_instance,
                                order=current_place
                            )
                            temp_img_data_pair = check_embedded_graphic(value)
                            if temp_img_data_pair is not None:

                                if field_exists('Answers', 'answer_graphic'):
                                    answer_instance.answer_graphic.save(temp_img_data_pair.actual_image_name,
                                                                        ContentFile(temp_img_data_pair.raw_image_data))
                                    answer_instance.save()

                                    # Parse the HTML using BeautifulSoup4 library
                                    html_obj = BeautifulSoup(value, 'html.parser')
                                    # Find the element with "img" tag
                                    my_img_element = html_obj.find('img')
                                    my_img_element['src'] = answer_instance.answer_graphic.url  # change src attribute
                                    value = str(html_obj)  # save html as string
                                    answer_instance.text = value  # update answer_instance text field
                                    answer_instance.save()  # save/update entry in database

                                if field_exists('Options', 'image'):
                                    # Parse the HTML using BeautifulSoup4 library
                                    html_obj = BeautifulSoup(value, 'html.parser')
                                    # Find the element with "img" tag
                                    my_img_element = html_obj.find('img')


                                    # save the image to ImageField
                                    option_instance.image.save(temp_img_data_pair.actual_image_name,
                                                               ContentFile(temp_img_data_pair.raw_image_data))
                                    my_img_element['src'] = option_instance.image.url

                                    value = str(html_obj)  # save html as string

                                    option_instance.text = value
                                    option_instance.save()

                        else:
                            options_instance = Options.objects.create(
                                text=value,
                                question=question_instance,
                                order=current_place
                            )
                            temp_img_data_pair = check_embedded_graphic(value)
                            if temp_img_data_pair is not None:
                                options_instance.image.save(temp_img_data_pair.actual_image_name,
                                                            ContentFile(temp_img_data_pair.raw_image_data))
                                options_instance.save()

                                # Parse the HTML using BeautifulSoup4 library
                                html_obj = BeautifulSoup(value, 'html.parser')
                                # Find the element with "img" tag
                                my_img_element = html_obj.find('img')
                                my_img_element['src'] = options_instance.image.url  # change src attribute
                                value = str(html_obj)  # save html as string
                                options_instance.text = value  # update option text field
                                options_instance.save()  # save/update entry in database

                elif the_question_type == 'matching_question':  # this is explicitly stated in rubric to support
                    # Canvas requires you to add at least one answer

                    the_question_type = 'ma'

                    answer_choices_dict = {}  # right side options and their ID's
                    left_side_dict = {}

                    question_instance = create_question(the_course, the_question_type, question_text_field,
                                                        max_points_for_question)
                    # this block of code checks for an embedded graphic in text, then saves it to field
                    # is there is a graphic. otherwise it does nothing.
                    image_data_pair = check_embedded_graphic(question_text_field)
                    if image_data_pair is not None:
                        # Save the image to the img field, then update the record/entry
                        question_instance.img.save(image_data_pair.actual_image_name,
                                                   ContentFile(image_data_pair.raw_image_data))
                        question_instance.save()
                        print(f"{question_instance.img.url}")

                        # Parse the HTML using BeautifulSoup4 library
                        html_obj = BeautifulSoup(question_text_field, 'html.parser')
                        # Find the element with "img" tag
                        my_img_element = html_obj.find('img')
                        my_img_element['src'] = question_instance.img.url  # change src attribute
                        question_text_field = str(html_obj)  # save html as string
                        question_instance.text = question_text_field  # update text field
                        question_instance.save()  # save/update entry in database

                    """
                    testquestion_instance = TestQuestion.objects.create(
                        test=test_instance,
                        question=question_instance,
                        assigned_points=max_points_for_question,
                        section=test_section_instance
                    )
                    """

                    ma_item_list.append({'question': question_instance, 'assigned_points': max_points_for_question})

                    # find left sides
                    for response_lid_elem in node.findall('response_lid'):  # for all response_lid elements in list
                        side_key = response_lid_elem.get('ident')
                        side_text = response_lid_elem.find('material').find('mattext').text
                        left_side_dict[side_key] = side_text
                    # find right sides
                    for response_label_elem in node.find('response_lid').find('render_choice').findall(
                            'response_label'):
                        side_key = response_label_elem.get('ident')
                        side_text = response_label_elem.find('.//mattext').text
                        answer_choices_dict[side_key] = side_text
                    # now find out which ones are matching pairs
                    matching_pairs_dict = {}
                    node = item.find('resprocessing')
                    right_side_key_to_delete_list = []
                    for respcondition_elem in node.findall('respcondition'):
                        varequal_elem = respcondition_elem.find('conditionvar').find('varequal')
                        if varequal_elem is not None:
                            left_key = varequal_elem.get('respident')
                            right_key = varequal_elem.text
                            right_side_key_to_delete_list.append(right_key)
                            # this makes a dictionary of matching pairs
                            matching_pairs_dict[left_side_dict.get(left_key)] = answer_choices_dict.get(right_key)
                    unique_key_list_to_del = list(
                        set(right_side_key_to_delete_list))  # this removes duplicate keys from list
                    for key_string in unique_key_list_to_del:
                        del answer_choices_dict[key_string]  # deletes a response option that was a correct right side

                    current_place = 0
                    pair_number = 0
                    # now save matching pairs to database
                    for key, value in matching_pairs_dict.items():
                        pair_number = pair_number + 1
                        current_place = current_place + 1
                        matching_pair_string = f"{{{key}, {value}}}"

                        # matching questions CANNOT have embedded graphics in responses
                        answer_instance = Answers.objects.create(
                            question=question_instance,
                            text=matching_pair_string
                        )

                        options_pair_dict = {'left': key, 'right': value, 'pairNum': pair_number}
                        option_instance = Options.objects.create(
                            question=question_instance,
                            order=current_place,
                            pair=options_pair_dict
                        )

                    current_place = 0
                    # save distractors to database
                    for value in answer_choices_dict.values():
                        current_place = current_place + 1
                        option_instance = Options.objects.create(
                            text=value,
                            question=question_instance,
                            order=current_place
                        )

                elif the_question_type == 'essay_question':
                    # mostly done but may need to process feedbacks or comments

                    the_question_type = 'es'

                    # this creates a question record in database
                    question_instance = create_question(the_course, the_question_type, question_text_field,
                                                        max_points_for_question)
                    # this block of code checks for an embedded graphic in text, then saves it to field
                    # is there is a graphic. otherwise it does nothing.
                    image_data_pair = check_embedded_graphic(question_text_field)
                    if image_data_pair is not None:
                        # Save the image to the img field, then update the record/entry
                        question_instance.img.save(image_data_pair.actual_image_name,
                                                   ContentFile(image_data_pair.raw_image_data))
                        question_instance.save()
                        print(f"{question_instance.img.url}")

                        # Parse the HTML using BeautifulSoup4 library
                        html_obj = BeautifulSoup(question_text_field, 'html.parser')
                        # Find the element with "img" tag
                        my_img_element = html_obj.find('img')
                        my_img_element['src'] = question_instance.img.url  # change src attribute
                        question_text_field = str(html_obj)  # save html as string
                        question_instance.text = question_text_field  # update text field
                        question_instance.save()  # save/update entry in database

                    """
                    testquestion_instance = TestQuestion.objects.create(
                        test=test_instance,
                        question=question_instance,
                        assigned_points=max_points_for_question,
                        section=test_section_instance
                    )
                    """

                    es_item_list.append({'question': question_instance, 'assigned_points': max_points_for_question})

                # commented out because currently not supported
                """
                elif the_question_type == 'fill_in_multiple_blanks_question':
                    print('')  # this is placeholder for code to extract info from question for table
                elif the_question_type == 'multiple_dropdowns_question':
                    print('')  # this is placeholder for code to extract info from question for table
                elif the_question_type == 'numerical_question':
                    print('')  # this is placeholder for code to extract info from question for table
                elif the_question_type == 'calculated_question':
                    print('')  # this is placeholder for code to extract info from question for table

                elif the_question_type == 'file_upload_question':
                    # should be done but may need to process feedbacks or comments
                    print('')
                elif the_question_type == 'text_only_question':
                    # placeholder for any future changes, but 99.9% sure this is done
                    print('')
                """

        # comment
        if mc_item_list:
            number_of_sections += 1
            test_section_instance = TestSection.objects.create(
                part=test_part_instance,
                section_number=number_of_sections,
                question_type='mc'
            )

            for item in mc_item_list:
                testquestion_instance = TestQuestion.objects.create(
                    test=test_instance,
                    question=item.get('question'),
                    assigned_points=item.get('assigned_points'),
                    section=test_section_instance
                )

                # update test section
                item.get('question').section = number_of_sections
                item.get('question').tests.add(test_instance)
                item.get('question').save()

        if tf_item_list:
            number_of_sections += 1
            test_section_instance = TestSection.objects.create(
                part=test_part_instance,
                section_number=number_of_sections,
                question_type='tf'
            )

            for item in tf_item_list:
                testquestion_instance = TestQuestion.objects.create(
                    test=test_instance,
                    question=item.get('question'),
                    assigned_points=item.get('assigned_points'),
                    section=test_section_instance
                )

                # update test section
                item.get('question').section = number_of_sections
                item.get('question').tests.add(test_instance)
                item.get('question').save()
                

        if fb_item_list:
            number_of_sections += 1
            test_section_instance = TestSection.objects.create(
                part=test_part_instance,
                section_number=number_of_sections,
                question_type='fb'
            )
            for item in fb_item_list:
                testquestion_instance = TestQuestion.objects.create(
                    test=test_instance,
                    question=item.get('question'),
                    assigned_points=item.get('assigned_points'),
                    section=test_section_instance
                )

                # update test section
                item.get('question').section = number_of_sections
                item.get('question').tests.add(test_instance)
                item.get('question').save()

        if ms_item_list:
            number_of_sections += 1
            test_section_instance = TestSection.objects.create(
                part=test_part_instance,
                section_number=number_of_sections,
                question_type='ms'
            )

            for item in ms_item_list:
                testquestion_instance = TestQuestion.objects.create(
                    test=test_instance,
                    question=item.get('question'),
                    assigned_points=item.get('assigned_points'),
                    section=test_section_instance
                )

                # update test section
                item.get('question').section = number_of_sections
                item.get('question').tests.add(test_instance)
                item.get('question').save()

        if es_item_list:
            number_of_sections += 1
            test_section_instance = TestSection.objects.create(
                part=test_part_instance,
                section_number=number_of_sections,
                question_type='es'
            )
            for item in es_item_list:
                testquestion_instance = TestQuestion.objects.create(
                    test=test_instance,
                    question=item.get('question'),
                    assigned_points=item.get('assigned_points'),
                    section=test_section_instance
                )

                # update test section
                item.get('question').section = number_of_sections
                item.get('question').tests.add(test_instance)
                item.get('question').save()

        if ma_item_list:
            number_of_sections += 1
            test_section_instance = TestSection.objects.create(
                part=test_part_instance,
                section_number=number_of_sections,
                question_type='ma'
            )

            for item in ma_item_list:
                testquestion_instance = TestQuestion.objects.create(
                    test=test_instance,
                    question=item.get('question'),
                    assigned_points=item.get('assigned_points'),
                    section=test_section_instance
                )

                # update test section
                item.get('question').section = number_of_sections
                item.get('question').tests.add(test_instance)
                item.get('question').save()

                #

    # file_info is just used for testing. remove after (probably)
    # for now, what the Parser returns depends on this
    file_info = None

    uploaded_file = None

    # """
    # 00 Begin
    # Code in triple quotes is used for when merged with frontend
    # "file" in request.FILES.get("file") changes or depends on something in the HTML/javascript form
    if request.method == "POST" and request.FILES.get("file"):
        uploaded_file = request.FILES["file"]  # Get the uploaded file
        print("File uploaded:", uploaded_file.name)

        file_info = {
            "filename": uploaded_file.name,
            "size": uploaded_file.size
        }
    else:
        print("No file uploaded to website.")

    if uploaded_file is None:
        return JsonResponse({"error": "No file uploaded or it doesn't exist.", "file_info": file_info}, status=400)

    course_id = request.POST.get("courseID")
    course_name = request.POST.get("courseName")
    course_crn = request.POST.get("courseCRN")
    course_semester = request.POST.get("courseSemester")

    course_instance, created = Course.objects.get_or_create(
        course_id=course_id,
        defaults={
            "name": course_name,
            "crn": course_crn,
            "sem": course_semester
        }
    )

    # Check if the user is authenticated (logged in)
    if request.user.is_authenticated:

        current_user = request.user

        course_instance.user = current_user  # sets field to current user
        course_instance.save()

        # Check if teacher already in course
        if current_user in course_instance.teachers.all():
            print(f'{current_user.username} teacher already in {course_instance.name} course')
        else:
            course_instance.teachers.add(current_user)  # Adds teacher if not in course
            course_instance.save()  # Updates the Course entry in the database (makes sure it's saved)
            print(f'{current_user.username} teacher ADDED to {course_instance.name} course')
    else:
        print("User is not logged in.")

    # Check if the user is authenticated (logged in)
    if request.user.is_authenticated:
        default_parsed_template, created = Template.objects.get_or_create(
            course=course_instance,
            name="QTI Default",
            author=request.user,
            defaults={
                # "author": request.user,
                "titleFont": "Times New Roman",
                "titleFontSize": 36,
                "subtitleFont": "Times New Roman",
                "subtitleFontSize": 24,
                "bodyFont": "Times New Roman",
                "bodyFontSize": 12,
                "pageNumbersInHeader": False,
                "pageNumbersInFooter": False,
                "partStructure": [{"sections": [{"questionType": "mc", "sectionNumber": 1},
                                                {"questionType": "tf", "sectionNumber": 2},
                                                {"questionType": "fb", "sectionNumber": 3},
                                                {"questionType": "es", "sectionNumber": 4},
                                                {"questionType": "ma", "sectionNumber": 5},
                                                {"questionType": "ms", "sectionNumber": 6}], "partNumber": 1}],
                "bonusSection": False,
                "published": True
            }
        )
        default_parsed_cpage, created = CoverPage.objects.get_or_create(
            course=course_instance,
            name="QTI Default",
            author=request.user,
            defaults={
            "testNum": 1,
            "date": date.today().isoformat(),
            "file": 'qti test',
            "showFilename": False,
            "blank": 'TR',
            "instructions": "Read all question instructions carefully!",
            "published": True
            }
        )
        default_parsed_template.coverPageID = default_parsed_cpage.id
        default_parsed_template.save()
    else:
        print("User is not authenticated")
        return JsonResponse({"error": "User is not authenticated."}, status=401)

    # 00 End
    # """

    # this is used to stop removing and adding "#" when switching between tests
    if uploaded_file is None:
        path_to_zip_file = 'qti sample w one quiz-slash-test w all typesofquestions.zip'
        # path_to_zip_file = 'added response feedback.zip'
        course_instance = Course.objects.first()
        if course_instance is None:
            course_instance = Course.objects.create(
                course_code='CS123',
                course_name='placeholder_course'
            )
    else:
        path_to_zip_file = uploaded_file

    if not path_to_zip_file.name.endswith('.zip'):
        return JsonResponse({"error": "File is not a QTI file."}, status=400)

    with zipfile.ZipFile(path_to_zip_file, 'r') as zip_ref:
        # List all files inside the zip file
        filename_list = zip_ref.namelist()

        if "imsmanifest.xml" not in filename_list:
            print("File is not valid QTI file.")
            return JsonResponse({"error": "File is not a valid QTI file."}, status=400)

        for file_name in filename_list:

            # looks for folders that are direct children of the zipfile
            if file_name.endswith('/'):

                temp_file_list = []

                # looks for files that are children of the found folder, then adds them to a list
                for temp_filename in filename_list:
                    if temp_filename.startswith(f'{file_name}') and (temp_filename != file_name):
                        temp_file_list.append(temp_filename)

                # if a folder is not empty, process the files in it with the parser
                if temp_file_list and len(temp_file_list) >= 2 and temp_file_list[0].endswith('.xml') and \
                        temp_file_list[1].endswith('.xml'):
                    # sort the files in the list because the metadata file for assessments is always
                    # named assessment_meta.xml, but the file with questions seems to always
                    # start with the letter "g"
                    temp_file_list = sorted(temp_file_list)
                    assessment_meta_path = temp_file_list[0]
                    questions_file_path = temp_file_list[1]

                    with zip_ref.open(assessment_meta_path) as outer_file:
                        with zip_ref.open(questions_file_path) as inner_file:
                            outer_file.seek(0)  # Reset file pointer
                            inner_file.seek(0)

                            # this calls the function that actually handles the parsing
                            parse_just_xml(outer_file, inner_file, course_instance, default_parsed_template)
                            #

    #
    end_time = time.perf_counter()
    execution_time = end_time - start_time
    print(execution_time)
    # this is here because the javascript that calls the Parser depends on what it returns
    if file_info is None:
        print("Successfully imported QTI file!")
        return JsonResponse({"success": "Successfully imported QTI file!"}, status=200)

    else:
        print("File processed successfully!")
        return JsonResponse({"success": "Successfully imported QTI file!", "file_info": file_info}, status=200)


#


def export_preview(request):
    if request.method == 'POST':

        type_of_preview = request.body.decode('utf-8')  # decode the byte string to normal string
        print(type_of_preview)

        courses = list(Course.objects.all().values('id', 'name'))

        tests = list(Test.objects.all().values('id', 'name'))

        questions = list(Question.objects.all().values('id', 'text', 'qtype'))
        options = list(Options.objects.all().values('id', 'text', 'question_id'))
        answers = list(Answers.objects.all().values('id', 'text', 'question_id'))

        return JsonResponse({
            'courses': courses,
            'tests': tests,
            'questions': questions,
            'options': options,
            'answers': answers
        })


def export_csv(request):
    #

    print("export_csv view triggered")  # used to make sure view function is being called

    # this function finds out whether the given record/entry is missing in the sheet by using record ID
    def is_record_missing(id_to_check, sheet):
        did_not_find = True
        column_of_cells = sheet['A']  # this grabs a column of cells from an Excel worksheet
        for cell in column_of_cells:
            if cell.value == id_to_check:
                did_not_find = False
                return did_not_find

        return did_not_find

    # returns dictionary of string-list pairs. each list is a list of IDs. uses ids_to_grab_list
    # ids_to_grab_list contains the column names of columns that need to be scraped from Excel sheet
    # lists_of_ids_dict is a dictionary of string-list pairs. each list has integers IDs
    def fill_out_sheet(sheet, query, lists_of_ids_dict, ids_to_grab_list):
        cursor = connection.cursor()
        cursor.execute(query)  # cursor holds result from cursor.execute() query
        # grabs all rows that cursor is holding. each list object is a tuple. doesn't grab column-name row
        rows = cursor.fetchall()
        # this extracts column names from table metadata
        # cursor.description is a list of tuples that contain table metadata
        # column_info designates the tuple object that contains column info
        # the first object in the column metadata-tuple is the column name; column_info[0]
        column_name_list = [column_info[0] for column_info in cursor.description]

        # this block enters the column names from database into the first row of Excel sheet
        # uses the form:        for index, item in enumerate(my_list, number):
        # number is where you start enumerating from. this enumeration starts from 1 (inclusive)
        column_list_pair_dict = {}  # BE CAREFUL! every column is actually column-1 for use with row list indexing
        for cell_column, column_name in enumerate(column_name_list, 1):
            cell_column_letter = get_column_letter(cell_column)  # get_column_letter maps number to letter
            sheet[f'{cell_column_letter}1'] = column_name

            if column_name in lists_of_ids_dict:
                row_index = cell_column - 1
                column_list_pair_dict[row_index] = lists_of_ids_dict.get(column_name)  # creates int-list pair

        current_records_in_file = 0
        print(column_list_pair_dict)
        for column_to_check, list_with_ids in column_list_pair_dict.items():
            # this block adds each desired database record to the Excel sheet
            # current_records_in_file = 0 # commented out because i might need to put it back later
            for row in rows:  # for every record/entry in list of records ...
                if row[column_to_check] is not None:
                    if row[column_to_check] in list_with_ids and is_record_missing(row[0], sheet):  # check ID in list
                        for cell_column, field_value in enumerate(row, 1):
                            cell_column_letter = get_column_letter(cell_column)
                            sheet[f'{cell_column_letter}{(current_records_in_file + 2)}'] = field_value
                        current_records_in_file += 1

        dict_to_return = None
        if ids_to_grab_list is not None:
            dict_to_return = {}  # dictionary of lists
            for index, column_id in enumerate(ids_to_grab_list, 0):
                desired_id = ids_to_grab_list[index]
                dict_to_return[desired_id] = []
                desired_id_column_letter = None

                # Find the column letter by matching the column header
                for first_row_cell in sheet[1]:
                    if first_row_cell.value == desired_id:
                        desired_id_column_letter = get_column_letter(first_row_cell.column)
                        break  # exit loop after finding match

                # Skip if column is not found
                if desired_id_column_letter is None:
                    print(f"[export_csv] Column '{desired_id}' not found in sheet '{sheet.title}'")
                    continue

                alleged_cell_column = sheet[desired_id_column_letter]
                for i, cell in enumerate(alleged_cell_column):
                    if i > 0 and cell.value is not None:  # skip header
                        dict_to_return[desired_id].append(cell.value)


        return dict_to_return

    def set_up_then_call():
        print(table_name)
        sheet = wb.create_sheet(title=f"{table_name}")
        query = f"SELECT * FROM `{table_name}`"  # this uses BACKTICKS to dynamically query the database

        needed_ids_list = copy.deepcopy(table_needs_these_ids_dict[table_name])
        id_dict = {}
        for id in needed_ids_list:
            if cache_of_ids_dict.get(id) is not None:  # avoid sending None objects instead of lists
                id_dict[id] = cache_of_ids_dict.get(id)  # creates a key-list pair
        id_dict = copy.deepcopy(id_dict)
        if table_name in need_to_use_regular_id_dict:
            if need_to_use_regular_id_dict.get(table_name) in id_dict:  # avoid accessing key-value that doesn't exist
                id_dict['id'] = id_dict.pop(
                    need_to_use_regular_id_dict.get(table_name))  # removes pair after getting value
        # id_dict['id'] = course_id_list  # this makes a key-list pair entry in a dictionary
        need_to_grab_ids_list = need_to_look_for_dict.get(table_name)  # gets list of ids to scrape from database

        # this if-block makes sure that we scrape the IDs if the records were not identified by ID
        if table_name in using_implied_id_dict and cache_of_ids_dict.get(using_implied_id_dict.get(table_name)) is None:
            if need_to_grab_ids_list is not None:
                need_to_grab_ids_list.append('id')
            else:
                need_to_grab_ids_list = ['id']

        result_dict = fill_out_sheet(sheet, query, id_dict, need_to_grab_ids_list)
        print(result_dict)
        if result_dict is not None:
            result_dict = copy.deepcopy(result_dict)  # make a copy of dict to avoid overwriting anything
            if 'id' in result_dict:
                result_dict[using_implied_id_dict.get(table_name)] = result_dict.pop(
                    'id')  # ex: replaces id with feedback_id
            for id_name, value in result_dict.items():
                if id_name not in cache_of_ids_dict and value:  # check if already have id list. also check if result value is not empty
                    cache_of_ids_dict[id_name] = value
        print('')  # just used to space things out for debugging

    print(request.method)

    if request.method == "POST":
        try:  # json.loads() will cause an error if the json is invalid or empty
            data = json.loads(request.body)  # parses json and creates/saves into a dictionary
        except json.JSONDecodeError:  # might change to "json.decoder.JSONDecodeError"
            print('Error. JSON is invalid or empty.')
            return JsonResponse({'error': 'Invalid or empty JSON provided'}, status=400)

        # gets lists from list dictionary. all of these are the ACTUAL IDs from the database (1st column)
        course_id_list = data.get('course', [])
        test_id_list = data.get('test', [])
        question_id_list = data.get('questions', [])
        type_of_export = data.get('typeOfExport', [])

        course_id_list = list(set(course_id_list))
        test_id_list = list(set(test_id_list))
        question_id_list = list(set(question_id_list))

        try:  # tries to convert every list into a list of integers (except for type_of_export)
            course_id_list = [int(course_id) for course_id in course_id_list]
            test_id_list = [int(test_id) for test_id in test_id_list]
            question_id_list = [int(question_id) for question_id in question_id_list]
        except ValueError:
            print('ID with NON-numeric ID-value provided')
            return JsonResponse({'error': 'ID with NON-numeric ID-value provided'}, status=400)

        try:
            export_type = type_of_export[0]
        except IndexError:
            print('Export type not given')
            return JsonResponse({'error': 'Export type not provided'}, status=400)

        """ this is to remind myself how to get a list of all the table names in the database (in case of a change)
        table_names = connection.introspection.table_names()
        print(table_names)
        """

        table_names_list = ['welcome_answers', 'welcome_attachment', 'welcome_course', 'welcome_coverpage',
                            'welcome_feedback', 'welcome_feedbackresponse',
                            'welcome_options', 'welcome_question', 'welcome_template', 'welcome_test',
                            'welcome_test_attachments', 'welcome_testpart', 'welcome_testquestion',
                            'welcome_testsection',
                            'welcome_textbook']

        course_table_names_list = ['welcome_course', 'welcome_textbook']

        test_table_names_list = ['welcome_test', 'welcome_template', 'welcome_coverpage', 'welcome_test_attachments',
                                 'welcome_attachment', 'welcome_testquestion', 'welcome_testsection',
                                 'welcome_testpart']

        question_table_names_list = ['welcome_question', 'welcome_options', 'welcome_answers', 'welcome_feedback',
                                     'welcome_feedbackresponse']

        # string-list dictionary where each list is a list of different ids needed when going through respective table
        table_needs_these_ids_dict = {
            'welcome_course': ['course_id'], 'welcome_textbook': ['textbook_id'],
            'welcome_test': ['test_id', 'course_id'], 'welcome_template': ['template_id', 'course_id'],
            'welcome_coverpage': ['coverPage', 'course_id'], 'welcome_test_attachments': ['test_id'],
            'welcome_attachment': ['attachment_id', 'course_id'], 'welcome_testquestion': ['test_id'],
            'welcome_testsection': ['section_id'], 'welcome_testpart': ['part_id', 'test_id'],
            'welcome_question': ['question_id', 'course_id'], 'welcome_options': ['question_id'],
            'welcome_answers': ['question_id'], 'welcome_feedback': ['question_id', 'test_id'],
            'welcome_feedbackresponse': ['feedback_id']
        }

        # string-list dictionary. when going through respective table, find ids in list
        need_to_look_for_dict = {
            'welcome_course': ['textbook_id'], 'welcome_textbook': None,
            'welcome_test': ['template_id'], 'welcome_template': ['coverPage'],
            'welcome_coverpage': None, 'welcome_test_attachments': ['attachment_id'],
            'welcome_attachment': None, 'welcome_testquestion': ['question_id', 'section_id'],
            'welcome_testsection': ['part_id'], 'welcome_testpart': None,
            'welcome_question': None, 'welcome_options': None,
            'welcome_answers': None, 'welcome_feedback': None,
            'welcome_feedbackresponse': None
        }

        # this dictionary is used like a boolean value. if the table name is a key, then it is seen as true.
        # if true, the value will get replaced with simply "id" when going through database
        need_to_use_regular_id_dict = {
            'welcome_course': 'course_id', 'welcome_textbook': 'textbook_id',
            'welcome_test': 'test_id', 'welcome_template': 'template_id',
            'welcome_coverpage': 'coverPage',
            'welcome_attachment': 'attachment_id',
            'welcome_testsection': 'section_id', 'welcome_testpart': 'part_id',
            'welcome_question': 'question_id'
        }

        # this holds the current ids acquired
        cache_of_ids_dict = {}

        # test_attachments, attachment,
        using_implied_id_dict = {
            'welcome_test': 'test_id', 'welcome_question': 'question_id', 'welcome_feedback': 'feedback_id'
        }

        wb = openpyxl.Workbook()  # creates the Workbook container object (Excel file)
        wb.remove(wb.active)  # Remove the default blank sheet

        list_of_valid_export_types = ['course', 'test', 'question', 'all courses', 'everything']

        if export_type not in list_of_valid_export_types:
            return JsonResponse({'error': 'Invalid export type'}, status=400)

        if export_type == 'all courses':
            print(export_type)  # placeholder

            query = f"SELECT * FROM welcome_course"
            cursor = connection.cursor()
            cursor.execute(query)  # cursor holds result from cursor.execute() query
            rows = cursor.fetchall()
            course_id_list = []
            for row in rows:
                if row[0] is not None:
                    course_id_list.append(row[0])
            if not course_id_list:
                print('No courses found')
                return JsonResponse({'not available': 'No courses found to export'}, status=400)
            export_type = 'course'

        if export_type == 'everything':
            entire_database_table_name_list = connection.introspection.table_names()
            print(entire_database_table_name_list)
            for table_name in entire_database_table_name_list:

                print(table_name)  # for debugging
                if len(table_name) > 31:
                    sheet = wb.create_sheet(title=f"{table_name[:31]}")
                else:
                    sheet = wb.create_sheet(title=f"{table_name}")
                query = f"SELECT * FROM `{table_name}`"  # this uses BACKTICKS to dynamically query the database
                cursor = connection.cursor()
                cursor.execute(query)  # cursor holds result from cursor.execute() query
                rows = cursor.fetchall()
                column_name_list = [column_info[0] for column_info in cursor.description]
                for cell_column, column_name in enumerate(column_name_list, 1):
                    cell_column_letter = get_column_letter(cell_column)  # get_column_letter maps number to letter
                    sheet[f'{cell_column_letter}1'] = column_name
                for row, whole_row in enumerate(rows, 2):
                    for column, field_value in enumerate(whole_row, 1):
                        column_letter = get_column_letter(column)  # columns are letters in Excel
                        sheet[f"{column_letter}{row}"] = field_value

        elif export_type == 'course':
            print(course_id_list)

            if not course_id_list:  # if given course list exists but is empty
                print('No courses given to export')
                return JsonResponse({'error': 'No courses given to export'}, status=400)

            cache_of_ids_dict['course_id'] = course_id_list

            for table_name in course_table_names_list + test_table_names_list + question_table_names_list:
                set_up_then_call()

        elif export_type == 'test':
            print(test_id_list)

            if not test_id_list:
                print('No tests given to export')
                return JsonResponse({'error': 'No tests given to export'}, status=400)

            cache_of_ids_dict['test_id'] = test_id_list

            for table_name in test_table_names_list + question_table_names_list:
                set_up_then_call()

        elif export_type == 'question':
            print(question_id_list)

            if not question_id_list:
                print('No questions given to export')
                return JsonResponse({'error': 'No questions given to export'}, status=400)

            cache_of_ids_dict['question_id'] = question_id_list

            for table_name in question_table_names_list:
                set_up_then_call()

        print(cache_of_ids_dict)
        print("Final sheets in workbook:", wb.sheetnames)

        # Save to a BytesIO stream instead of a file
        from io import BytesIO
        output = BytesIO()
        wb.save(output)  # this will CRASH the program if there are no Excel sheets in the workbook
        output.seek(0)

        # Prepare the response
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=exported_data.xlsx'

        print("did it get here???")  # used for debugging

        return response

    return JsonResponse({'error': 'request.method was not POST'}, status=400)


#



def import_csv(request):

    # start of import_csv function
    print("import_csv views function started")

    imported_csv_file = None

    if request.method != "POST":
        print("request.method was not POST")
        return JsonResponse({'error': 'request.method was not POST', 'success': False}, status=400)

    if request.FILES.get("imported_csv_file"):
        imported_csv_file = request.FILES["imported_csv_file"]  # Get the uploaded file
        print("File uploaded:", imported_csv_file.name)
    else:
        print("No file uploaded to website.")
        return JsonResponse({'error': 'No file uploaded to website.', 'success': False}, status=400)

    if  not imported_csv_file.name.endswith(".csv"):
        print("File uploaded is not a .csv file.")
        return JsonResponse({'error': 'File uploaded is not a .csv file.', 'success': False}, status=400)

    chosen_table_name = None
    if request.POST.get("table_to_import"):
        chosen_table_name = request.POST.get("table_to_import")
    else:
        print("No table name was given to import")
        return JsonResponse({'error': 'No table name was given to import', 'success': False}, status=400)

    delete_all_rows = request.POST.get("deleteAllRows")

    if delete_all_rows is None:
        print("No indication as to whether or not to delete all rows was given")
        return JsonResponse({'error': 'No indication as to whether or not to delete all rows was given', 'success': False}, status=400)

    should_del_all_rows = False
    if delete_all_rows == "true":
        should_del_all_rows = True

    try:
        # [possibly detect encoding in file. ex: utf-8]
        # encoding only supports utf-8

        raw_file_bytes = imported_csv_file.read() # files uploaded by HTTP are always bytes. bytes object
        encoding_type = 'utf-8'
        decoded_file_bytes = raw_file_bytes.decode(encoding_type) # turn raw bytes into a Python string
        file_like_string = io.StringIO(decoded_file_bytes) # wraps the string in a file-like object

        # df is a Pandas DataFrame. pd is "alias" for Pandas library
        df = pd.read_csv(file_like_string)

    except pd.errors.ParserError: # if csv-parsing error
        print("Error while parsing CSV. The uploaded file may not be a valid CSV file.")
        return JsonResponse({'error': 'Error while parsing CSV. The uploaded file may not be a valid CSV file.', 'success': False}, status=400)
    except UnicodeDecodeError: # if decoding error
        print("The uploaded file could not be decoded")
        return JsonResponse({'error': 'The uploaded file could not be decoded', 'success': False}, status=400)
    except Exception as e:
        print(f"An unexpected error occurred in Django while initially handling the file: {str(e)}")
        return JsonResponse({'error': 'An unexpected error occurred while initially handling the file.', 'success': False}, status=400)

    if df.empty: # if csv is empty
        print("The uploaded CSV file is empty")
        return JsonResponse({'error': 'The uploaded CSV file is empty', 'success': False}, status=400)

    model_name = chosen_table_name

    try:
        model = apps.get_model(app_label='welcome', model_name=model_name) # get Model class
        print("Importing into model:", model.__name__)
    except LookupError: # if model could not be found
        return JsonResponse({'error': f'Model {model_name} not found.', 'success': False}, status=400)

    model.objects.all().delete() # deletes all rows in respective table

    rows_skipped = [] # keeps track of rows that couldn't be processed

    model_fields = {field.name: field for field in model._meta.fields} # model._meta.fields is a list of field objects

    # this code checks to make sure every given column is a field in Model. returns if a column is not a real field
    column_names_list = list(df.columns)
    for column_name in column_names_list:
        if column_name not in model_fields:
            if column_name.endswith('_id'):
                real_name = column_name[:-3]
                if real_name not in model_fields:
                    print(f"Error: Column name {real_name} not found in model fields.")
                    return JsonResponse({'error': f"Column name {real_name} not found in model fields.", 'success': False}, status=400)
            else:
                print(f"Error: Column name {column_name} not found in model fields.")
                return JsonResponse({'error': f"Column name {column_name} not found in model fields.", 'success': False}, status=400)

    # for debugging

    # print(model_fields)
    # for field in model_fields:
    #     print(field)

    successfully_imported_rows = 0

    for index, row in df.iterrows():
        should_continue = False
        new_record_data_dict = {}

        #print(index+1) # debugging

        for column_name, value in row.items():

            if column_name.endswith('_id'):
                if model_name == 'Course' and column_name == 'course_id':
                    field = model_fields.get(column_name)
                else:
                    real_name = column_name[:-3]
                    field = model_fields.get(real_name)
            else:
                field = model_fields.get(column_name)

            if field is None: # if field with column name doesn't exist
                rows_skipped.append(int(index)+1)
                should_continue = True
                break

            if isinstance(value, str) and value == 'NULL':
                value = None

            if value is None or pd.isna(value): # NaN is pandas version of NULL or None. one result of pd.isna() is whether value is NaN or not; boolean
                if not field.null:  # if field does NOT allow a null value
                    rows_skipped.append(int(index)+1)
                    should_continue = True
                    break
                else:  # if field allows a null value
                    new_record_data_dict[field.name] = None
            else:

                if field.is_relation:  # check if ForeignKey

                    related_model = field.related_model # get Model class
                    try:
                        related_instance = related_model.objects.get(id=int(value)) # get instance of related Model
                        new_record_data_dict[field.name] = related_instance
                    except ValueError:
                        print('Non integer ID provided for ForeignKey')
                        rows_skipped.append(int(index) + 1)
                        should_continue = True
                        break
                    except related_model.DoesNotExist:  # if the desired instance doesn't exist
                        print(
                            f"Record with ID of {value} does not exist in database table for {related_model.__name__} model")
                        rows_skipped.append(int(index)+1)
                        should_continue = True
                        break

                else:
                    if isinstance(field, IntegerField):
                        try:
                            value = int(value)
                        except ValueError:
                            rows_skipped.append(int(index) + 1)
                            should_continue = True
                            break
                    elif isinstance(field, FloatField):
                        try:
                            value = float(value)
                        except ValueError:
                            rows_skipped.append(int(index) + 1)
                            should_continue = True
                            break
                    elif isinstance(field, BooleanField):
                        # Convert strings like '1', '0', 'true', 'false'
                        if str(value).strip().lower() in ['1', 'true']:
                            value = True
                        elif str(value).strip().lower() in ['0', 'false']:
                            value = False
                        else:
                            rows_skipped.append(int(index) + 1)
                            should_continue = True
                            break

                    new_record_data_dict[field.name] = value


        if should_continue:
            continue # probably redundant. skips to next iteration. stops executing this iteration.
        else:
            try:
                model.objects.create(**new_record_data_dict)  # create record in database
                successfully_imported_rows += 1
            except IntegrityError:
                print(f"Record with given ID already exists in database table for model")
                rows_skipped.append(int(index) + 1)

    if successfully_imported_rows == 0:
        print('No rows were imported')
        return JsonResponse({'error': "All rows were skipped", 'success': False}, status=400)
    if rows_skipped:
        print(f"rows skipped: {rows_skipped}")
        return JsonResponse({'message': 'CSV file imported. Some rows were skipped.', 'success': True, 'rows_skipped': rows_skipped}, status=200)
    else:
        print("All rows were imported successfully!")
        return JsonResponse({'message': 'CSV file imported. All rows were imported successfully!', 'success': True}, status=200)

    #print("end of import_csv function.")
    # end of import_csv function
#


def create_csv_template(request):

    # start of create_csv_template function
    print('start of create_csv_template function')

    model_for_template = request.body.decode('utf-8')  # decode the byte string to normal string
    print(f"Model for template: {model_for_template}")
    if model_for_template is None or model_for_template == '':
        return JsonResponse({'error': 'No model was given to create a CSV template for'}, status=400)

    column_names_list = []
    row_one = []
    row_two = []

    if model_for_template == 'Textbook':
        column_names_list = ['id', 'title', 'author', 'version', 'isbn', 'link', 'published', 'publisher_id']
        row_one = ['1', '"Linear Algebra"', '"Math Teacher"', '5', '123456789', 'amazon.com/book1', '1', '5']
        row_two = ['2', '"C++ for Everyone"', '"Some other author"', '3', '678216378', 'pearson.com/book2', '1', '3']
    elif model_for_template == 'UserProfile':
        column_names_list = ['id', 'role', 'user_id']
        row_one = ['1', 'teacher', '1']
        row_two = ['2', 'publisher', '2']
    elif model_for_template == 'Course':
        column_names_list = ['id', 'course_id', 'name', 'crn', 'sem', 'user_id', 'published']
        row_one = ['1', 'CS499' ,'"Senior Project Design"', '3213', '"Fall 2024"', '1', '1']
        row_two = ['2', 'MA385' ,'"Intro to Probability and Statistics"', '2131', '"Spring 2025"', '1', '1']
    elif model_for_template == 'Question':
        column_names_list = ['id', 'text', 'ansimg', 'score', 'eta', 'reference', 'directions', 'comments', 'created_at', 'updated_at', 'course_id', 'author_id', 'chapter', 'answer', 'section', 'img', 'qtype', 'textbook_id', 'published']
        row_one = ['1', '"Who is the greatest Fortnite player in the world?"', '"/media/ansimg1.jpg"', '2', '5', 'reference1', 'directions1', '"sample comment"', '"2025-04-16 10:32:03.405280"', '"2025-04-16 10:32:03.390905"', '1', '1', '0', 'NULL', '1', 'media/img1.jpg', 'mc', '4', '1']
        row_two = ['2', '"How much wood would a ..."', '"/media/ansimg1.jpg"', '4', '2', 'reference2', 'directions2', '"another sample comment"', '"2025-04-17 10:32:03.385875"', '"2025-04-27 10:32:03.405297"', '2', '1', '1', 'NULL', '2', 'media.img2.jpg', 'fb', '2', '0']
    elif model_for_template == 'Options':
        column_names_list = ['id', 'text', 'question_id', 'image', 'order', 'pair']
        row_one = ['1', '"yes because it is needed"', '3', 'NULL', '1', 'NULL']
        row_two = ['2', '"Certain activities are prohibited."', '2', '/media/image.jpg', '2', 'NULL']
    elif model_for_template == 'Answers':
        column_names_list = ['id', 'text', 'answer_graphic', 'response_feedback_text', 'response_feedback_graphic', 'question_id', 'pair']
        row_one = ['1', 'D', 'NULL', 'NULL', '', '1', 'NULL']
        row_two = ['2', 'true', '', 'NULL', '/media/feedback_graphic.jpg', '2', 'NULL']
    elif model_for_template == 'Template':
        column_names_list = ['id', 'name', 'bodyFont', 'bodyFontSize', 'headerText', 'footerText', 'course_id', 'coverPageID', 'pageNumbersInFooter', 'pageNumbersInHeader', 'subtitleFont', 'subtitleFontSize', 'titleFont', 'titleFontSize', 'textbook_id', 'partStructure', 'bonusSection', 'published', 'author_id', 'courseTag', 'dateTag', 'nameTag', 'bonusQuestions']
        row_one = ['1', '"System Default"', '"Times New Roman"', '12', '', '"Please read all questions carefully"', '1', '0', '1', '0', '"Times New Roman"', '24', '"Times New Roman"', '36', 'NULL', '"parts and sections"', '0', '1', '1', '', '', '', '[]']
        row_two = ['2', '"QTI Default"', '"Times New Roman"', '12', 'NULL', 'NULL', '1', '0', '0', '0', '"Times New Roman"', '24', '"Times New Roman"', '36', '1', '"parts and section"', '0', '1', '1', 'NULL', 'NULL', 'NULL', 'NULL']
    elif model_for_template == 'CoverPage':
        column_names_list = ['id', 'name', 'testNum', 'date', 'file', 'showFilename', 'instructions', 'course_id', 'blank', 'published', 'textbook_id', 'author_id']
        row_one = ['1', 'nameOne', '1', '"2025-04-19"', 'defaultpage', '1', '"Grade according to the rubric giving partial credit where indicated"', '1', 'TR', '1', 'NULL', '1']
        row_two = ['2', '"nameTwo"', '2', '"2025-04-19"', '"defaultpage_3"', '1', '"Grade according to the rubric giving partial credit where indicated"', '1', 'TR', '1', 'NULL', '1']
    elif model_for_template == 'Attachment':
        column_names_list = ['id', 'file', 'course_id', 'name', 'published', 'textbook_id', 'author_id']
        row_one = ['1', '"attachments/example.csv"', '1', '"attachment name 1"', '0', 'NULL', '1']
        row_two = ['2', '"attachments/another_example.jpg"', '2', 'name2', '0', 'NULL', '1']
    elif model_for_template == 'Test':
        column_names_list = ['id', 'name', 'date', 'filename', 'is_final', 'created_at', 'updated_at', 'course_id', 'template_id', 'textbook_id', 'templateID', 'author_id']
        row_one = ['1', '"Final Exam"', 'NULL', 'NULL', '1', '"2025-04-17 10:32:03.385875"', '"2025-04-17 10:32:03.385875"', '3', '7', '6', '7', '2']
        row_two = ['2', '"Test of Honor"', 'NULL', 'NULL', '0', '"2025-04-17 10:32:03.385875"', '"2025-04-17 10:32:03.385875"', '3', '11', '2', '11', '1']
    elif model_for_template == 'TestPart':
        column_names_list = ['id', 'part_number', 'test_id']
        row_one = ['1', '1', '1']
        row_two = ['2', '2', '1']
    elif model_for_template == 'TestSection':
        column_names_list = ['id', 'section_number', 'question_type', 'part_id']
        row_one = ['1', '1', 'mc', '1']
        row_two = ['2', '1', 'tf', '2']
    elif model_for_template == 'TestQuestion':
        column_names_list = ['id', 'assigned_points', 'order', 'randomize', 'special_instructions', 'question_id', 'test_id', 'section_id']
        row_one = ['1', '5', '1', '1', '"special instructions"', '2', '6', '1']
        row_two = ['2', '7', '2', '0', '"more special instructions"', '5', '3', '2']
    elif model_for_template == 'Feedback':
        column_names_list = ['id', 'rating', 'comments', 'created_at', 'user_id', 'question_id', 'test_id', 'averageScore', 'time']
        row_one = ['1', '5', 'comment1', '"2025-04-17 10:32:03.385875"', '5', '4', '3', '"93.24"', '']
        row_two = ['2', '2', '"comment 2"', '"2025-04-17 10:32:03.385875"', '1', '5', '1', '"77.01"', 'NULL']
    elif model_for_template == 'FeedbackResponse':
        column_names_list = ['id', 'text', 'date', 'created_at', 'feedback_id', 'user_id']
        row_one = ['1', '"feedback response"', '"2025-04-17 10:32:03.385875"', '"2025-04-17 10:32:03.385875"', '3', '1']
        row_two = ['2', '"yes morefeedbackreponses"', '"2025-04-17 10:32:03.385875"', '"2025-04-17 10:32:03.385875"', '2', '4']
    else:
        print('Invalid model name given')
        return JsonResponse({'error': 'Invalid model name given'}, status=400)


    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="import_template.csv"'

    #file_writer = csv.writer(response)
    file_writer = csv.writer(response, quoting=csv.QUOTE_MINIMAL, quotechar="'")

    file_writer.writerow(column_names_list)
    file_writer.writerow(row_one)
    file_writer.writerow(row_two)

    print('Template csv created!')
    return response

    # end of create_csv_template function
    # print('end of create_csv_template function')
#


# testing connecting the html files


#each of these functions renders a different html file

def login_view(request):
    return render(request, 'welcome/login.html')

def signup_view(request):
    return render(request, 'welcome/signup.html')

@login_required
def teacher_dashboard(request):
    context = {'username': request.user.username, 'userRole': "teacher"}
    return render(request, 'welcome/SBteacher.html', context)


@login_required
def publisher_dashboard(request):
    context = {'username': request.user.username, 'userRole': "publisher"}
    return render(request, 'welcome/SBpublisher.html', context)

@login_required
def webmaster_dashboard(request):
    context = {'username': request.user.username, 'userRole': "webmaster"}
    return render(request, 'welcome/webmaster.html', context)


"""
Beginning to test the frontend to the backend connection
"""

# function to handle the signup form
def signup_handler(request):

    # 
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        passwordconfirm = request.POST.get('passwordconfirm')
        role = request.POST.get('role')
        
        # Check if passwords match
        if password != passwordconfirm:
            messages.error(request, "Passwords do not match.")
            return redirect('signup')  # Assumes you have a URL named 'signup' for the signup page

        try:
            # Create the new user
            user = User.objects.create_user(username=username, password=password)
            user.save()
            
            # Create the user profile with the role chosen
            profile = UserProfile.objects.create(user=user, role=role)
            profile.save()
            return redirect('home')  # Redirect to the home page or another page of your choice
        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
            return redirect('signup')
    else:
        # For a GET request, simply render the signup form
        return render(request, 'signup.html')




"""
FAQ
"""
def faq_view(request):
    return render(request, 'welcome/faq.html')